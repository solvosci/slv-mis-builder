# © 2025 Solvos Consultoría Informática (<http://www.solvos.es>)
# License LGPL-3 - See http://www.gnu.org/licenses/lgpl-3.0.html

from odoo import models, fields, _
from odoo.osv import expression
from datetime import timedelta
import openpyxl.utils


class ProjectMisXlsxReport(models.AbstractModel):
    _name = 'report.project_mis_xlsx_report.project_mis'
    _inherit = 'report.report_xlsx.abstract'
    _description = "XLSX model report"

    def get_months_between_dates(self, start_date, end_date):
        months = []
        if start_date > end_date:
            start_date, end_date = end_date, start_date

        current_date = fields.Date.to_date(f"{start_date.year}-{start_date.month:02d}-01")
        end_date = fields.Date.to_date(f"{end_date.year}-{end_date.month:02d}-01")

        while current_date <= end_date:
            months.append(current_date)
            year = current_date.year + (current_date.month // 12)
            month = (current_date.month % 12) + 1
            current_date = fields.Date.to_date(f"{year}-{month:02d}-01")
        return months

    def _get_amount(self, model, domain, field='amount', negate=False):
        result = sum(self.env[model].search(domain).mapped(field))
        return -result if negate else result

    def _get_cost_by_account_codes(self, analytic, first_day, last_day, account_codes):
        base_domain = [
            ('analytic_account_id', '=', analytic.id),
            ('date', '>=', first_day),
            ('date', '<=', last_day),
        ]
        if account_codes:
            account_domain = expression.OR([[('account_id.code', 'like', code)] for code in account_codes])
            base_domain = expression.AND([base_domain, account_domain])

        lines = self.env['account.move.line'].search(base_domain)
        return sum(lines.mapped('debit')) - sum(lines.mapped('credit'))

    def _get_credit_by_account_codes(self, analytic, first_day, last_day, account_codes):
        base_domain = [
            ('analytic_account_id', '=', analytic.id),
            ('date', '>=', first_day),
            ('date', '<=', last_day),
        ]
        if account_codes:
            account_domain = expression.OR([[('account_id.code', 'like', code)] for code in account_codes])
            base_domain = expression.AND([base_domain, account_domain])

        lines = self.env['account.move.line'].search(base_domain)
        return sum(lines.mapped('credit'))

    def get_labor_cost(self, analytic, first_day, last_day):
        domain = [
            ('account_id', '=', analytic.id),
            ('date', '>=', first_day),
            ('date', '<=', last_day),
            ('project_id', '!=', False),
        ]
        return self._get_amount('account.analytic.line', domain, 'amount', negate=True)

    def get_material_cost(self, analytic, first_day, last_day):
        return self._get_cost_by_account_codes(analytic, first_day, last_day, ['600000%', '602000%']) - self._get_credit_by_account_codes(analytic, first_day, last_day, ['778000%'])

    def get_subcontracting_cost(self, analytic, first_day, last_day):
        return self._get_cost_by_account_codes(analytic, first_day, last_day, ['607000%', '607002%', '623009%', '623010%', '623014%'])

    def get_license_cost(self, analytic, first_day, last_day):
        return self._get_cost_by_account_codes(analytic, first_day, last_day, ['621000%', '621006%'])

    def get_travel_and_other_cost(self, analytic, first_day, last_day):
        return self._get_cost_by_account_codes(analytic, first_day, last_day, [
            '624000%', '629009%', '629015%', '640002%', '640003%', '649001%'
        ])

    def get_inventory_variation_cost(self, analytic, first_day, last_day):
        domain = [
            ('analytic_account_id', '=', analytic.id),
            ('date', '>=', first_day),
            ('date', '<=', last_day),
            ('account_id.code', 'like', '610000%'),
            ('analytic_mrp_from_child', '!=', True),
            ('analytic_mrp_unbuilt', '!=', True),
        ]
        return self._get_amount('account.move.line', domain, 'debit')

    def _get_real_order(self, analytic, first_day, last_day, excluded_products=False, include_products=False):
        domain = [
            ('account_analytic_id', '=', analytic.id),
            ('forecast_date', '>=', first_day),
            ('forecast_date', '<=', last_day),
            ('state', 'not in', ['draft', 'cancel', 'sent', 'to approve']),
            ('qty_invoiced', '=', 0),
        ]
        if excluded_products:
            domain.append(('product_id.default_code', 'not in', excluded_products))
        if include_products:
            domain.append(('product_id.default_code', 'in', include_products))

        lines = self.env['purchase.order.line'].search(domain)
        subtotal = sum(lines.mapped('price_subtotal'))
        qty = sum(lines.mapped('product_qty'))
        if not subtotal or not qty:
            return 0
        return (subtotal / qty) * (qty - sum(lines.mapped('qty_invoiced')))

    def get_real_material_order(self, analytic, first_day, last_day):
        return self._get_real_order(analytic, first_day, last_day, excluded_products=['S1017', '70617-001'])

    def get_real_services_order(self, analytic, first_day, last_day):
        return self._get_real_order(analytic, first_day, last_day, include_products=['S1017'])

    def get_real_travel_order(self, analytic, first_day, last_day):
        return self._get_real_order(analytic, first_day, last_day, include_products=['70617-001'])

    def get_expenses(self, analytic, first_day, last_day):
        return (
            self.get_labor_cost(analytic, first_day, last_day) +
            self.get_material_cost(analytic, first_day, last_day) +
            self.get_subcontracting_cost(analytic, first_day, last_day) +
            self.get_license_cost(analytic, first_day, last_day) +
            self.get_travel_and_other_cost(analytic, first_day, last_day) +
            self.get_inventory_variation_cost(analytic, first_day, last_day)
        )

    def get_billing(self, analytic, first_day, last_day):
        return self._get_cost_by_account_codes(analytic, first_day, last_day, [
            '700000%', '705000%', '710000%', '733000%'
        ]) * -1

    def get_magin(self, analytic, first_day, last_day):
        domain = [
            ('analytic_account_id', '=', analytic.id),
            ('date_from', '>=', first_day),
            ('date_from', '<=', last_day),
            ('kpi_expression_id.kpi_id', 'like', 'margen'),
            ('budget_id.state', '=', 'confirmed'),
        ]
        return self._get_amount('mis.budget.item', domain)

    def get_income(self, analytic, first_day, last_day):
        margin = self.get_magin(analytic, first_day, last_day)
        if not margin or margin < -100:
            return 0
        expenses = self.get_expenses(analytic, first_day, last_day)
        return expenses / (1 - (margin / 100))

    def get_result(self, analytic, first_day, last_day):
        income = self.get_income(analytic, first_day, last_day)
        expenses = self.get_expenses(analytic, first_day, last_day)
        result = income - expenses
        return result if result > 0 else 0

    def get_accumulated_year(self, row, col, month, months):
        if month.month == 1:
            cell = openpyxl.utils.get_column_letter(col + 1)
            return f"=SUM({cell}{row})"
        start_letter = openpyxl.utils.get_column_letter(col)
        end_letter = openpyxl.utils.get_column_letter(col + 1)
        return f"=SUM({start_letter}{row + 1},{end_letter}{row})"

    def get_accumulated_total(self, row, col, month, months):
        start_letter = openpyxl.utils.get_column_letter(col)
        end_letter = openpyxl.utils.get_column_letter(col + 1)
        return f"=SUM({start_letter}{row + 1},{end_letter}{row - 1})"

    def add_row(self, worksheet, row, col, name, datas, analytic, style=None):
        months = self.get_months_between_dates(datas.date_from, datas.date_to)
        for month in months:
            col += 1
            first_day = month
            last_day = (month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            total = 0

            if name == _('Result'):
                total = self.get_result(analytic, first_day, last_day)
            elif name == _('Income'):
                total = self.get_income(analytic, first_day, last_day)
            elif name == _('Billing'):
                total = self.get_billing(analytic, first_day, last_day)
            elif name == _('Expense'):
                total = self.get_expenses(analytic, first_day, last_day)
            elif name == _('Labor'):
                total = self.get_labor_cost(analytic, first_day, last_day)
            elif name == _('Materials'):
                total = self.get_material_cost(analytic, first_day, last_day)
            elif name == _('Subcontracting'):
                total = self.get_subcontracting_cost(analytic, first_day, last_day)
            elif name == _('Licenses'):
                total = self.get_license_cost(analytic, first_day, last_day)
            elif name == _('Travel and Other'):
                total = self.get_travel_and_other_cost(analytic, first_day, last_day)
            elif name == _('Inventory Change'):
                total = self.get_inventory_variation_cost(analytic, first_day, last_day)

            if _('Yearly Cumulate') in name:
                formula = self.get_accumulated_year(row, col, month, months)
                worksheet.write_formula(row, col, formula, style)
            elif _('Total Cumulate') in name:
                formula = self.get_accumulated_total(row, col, month, months)
                worksheet.write_formula(row, col, formula, style)
            else:
                worksheet.write(row, col, total, style)

        col += 1
        if _('Yearly Cumulate') not in name and _('Total Cumulate') not in name:
            start_letter = openpyxl.utils.get_column_letter(2)
            end_letter = openpyxl.utils.get_column_letter(col)
            formula = f"=SUM({start_letter}{row + 1}:{end_letter}{row + 1})"
            worksheet.write_formula(row, col, formula, style)
        else:
            worksheet.write(row, col, "", style)

    def generate_xlsx_report(self, workbook, data, datas):
        worksheet = workbook.add_worksheet(_('Sheet 1'))
        account_analytic_ids = datas.account_analytic_ids

        worksheet.set_column('A:A', 16)
        worksheet.set_column('B:ZZ', 9)

        analytic_format = workbook.add_format({'bold': True, 'bg_color': '#0A2F7F', 'color': 'white', 'font_size': 10})
        base_format = workbook.add_format({'font_size': 8, 'num_format': '#,##0.00'})
        date_format = workbook.add_format({'font_size': 8, 'bold': True, 'align': 'center', 'bg_color': '#F0EEEE'})
        income_format = workbook.add_format({'bg_color': '#BCF4A9', 'font_size': 8, 'num_format': '#,##0.00'})
        expense_format = workbook.add_format({'bg_color': '#FBA7A7', 'font_size': 8, 'num_format': '#,##0.00'})

        row_fields = [
            _('Result'), _('Income'), _('Yearly Cumulate Income'), _('Total Cumulate Income'),
            _('Billing'), _('Expense'), _('Yearly Cumulate Expense'), _('Total Cumulate Expense'),
            _('Labor'), _('Materials'), _('Subcontracting'), _('Licenses'),
            _('Travel and Other'), _('Inventory Change')
        ]

        row, col = 0, 0
        months = self.get_months_between_dates(datas.date_from, datas.date_to)

        # Header
        for month in months:
            col += 1
            worksheet.write(row, col, month.strftime('%m/%Y'), date_format)
        worksheet.write(row, col + 1, 'Total', date_format)

        # Totals
        row += 1
        worksheet.merge_range(row, 0, row, col + 1, _('TOTALS'), analytic_format)
        for field in row_fields:
            row += 1
            col = 0
            style = income_format if _('Income') in field else expense_format if _('Expense') in field else base_format
            worksheet.write(row, col, field, style)
            for i, month in enumerate(months):
                column_letter = openpyxl.utils.get_column_letter(i + 2)
                sum_fields = [f"{column_letter}{(row + 15 * (j + 1) + 1)}" for j in range(len(account_analytic_ids))]
                formula = f"=SUM({','.join(sum_fields)})"
                worksheet.write_formula(row, i + 1, formula, style)

            # Total line
            col += 1
            if _('Yearly Cumulate') not in field and _('Total Cumulate') not in field:
                start_letter = openpyxl.utils.get_column_letter(2)
                end_letter = openpyxl.utils.get_column_letter(i + 2)
                formula = f"=SUM({start_letter}{row + 1}:{end_letter}{row + 1})"
                worksheet.write_formula(row, i + 2, formula, style)
            else:
                worksheet.write(row, i + 2, "", style)

        # Detalle por cuenta analítica
        row = 15
        for analytic in account_analytic_ids:
            row += 1
            worksheet.merge_range(row, 0, row, len(months) + 1, analytic.name, analytic_format)

            for field in row_fields:
                row += 1
                style = income_format if _('Income') in field else expense_format if _('Expense') in field else base_format
                worksheet.write(row, 0, field, style)
                self.add_row(worksheet, row, 0, field, datas, analytic, style)
